"""
Generate the final Excel file with all persons mentioned in
'Ni venganza ni perdón' by Julio Scherer Ibarra,
ranked by Scherer's opinion of them.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DIR = r"c:\Users\User\Documents\PROYECTOS_PERSONALES\scherer\output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =============================================================================
# CONSOLIDATED DATA from all 15 chunks (pages 1-300)
# Each entry: (Name, Opinion_word, Score, Description, Pages, Weight)
#
# Opinion scale:
# Desprecio(-5), Hostilidad(-4), Crítica(-3), Desconfianza(-2),
# Indiferencia(-1), Neutral(0), Cordialidad(1), Aprecio(2),
# Admiración(3), Lealtad(4), Devoción(5)
#
# Weight (Peso narrativo): based on pages dedicated to the person
# Alta = 15+ pages, Media = 5-14 pages, Baja = 1-4 pages
# =============================================================================

personas = [
    # --- MUY NEGATIVOS (-5 a -4) ---
    ("Francisco Gil Díaz", "Desprecio", -5,
     "Lo describe como 'malo, malísimo', que ejercía 'terrorismo fiscal'. Lo persiguió durante 5 años con 17 denuncias, intentó chantajearlo para controlar Proceso. En la UP le dijo que 'no tenía el alma podrida'.",
     "27-32, 203, 263-264, 284"),

    ("Jesús Ramírez Cuevas", "Desprecio", -5,
     "Lo califica de 'gran manipulador', 'personaje espantoso' que 'hizo mucho daño al Gobierno'. Lo acusa de operar la mañanera a su conveniencia, filtrar rumores, crear estructura mediática ilegítima, y reunirse con 'el rey del huachicol'. Asesoró la desastrosa defensa del T-MEC sobre glifosato y maíz transgénico sin tener formación legal, provocando un 'revés histórico para la soberanía alimentaria'. Convenció al presidente junto con Álvarez-Buylla de prohibir el glifosato sin sustituto, dañando gravemente la agricultura. Su nombre aparece en documentos judiciales de EE.UU.",
     "89, 127, 197-200, 201-221, 249, 271, 284"),

    ("Alejandro Gertz Manero", "Desprecio", -5,
     "Lo acusa de estar 'detrás de la persecución jurídica' en su contra, de actuar 'como extorsionador', de fabricar escándalos fallidos desde la FGR. Le dijo 'usted optó por tener fiscal enemigo'. Compara su lógica de chantaje con la de Gil Díaz. La confrontación con el Fiscal es uno de los hilos conductores de todo el libro.",
     "28, 108, 128-129, 253-258, 262, 276, 284-285"),

    ("Hernán Gómez Bruera", "Desprecio", -5,
     "Autor de un 'libro impúdico' y 'miserable', 'plagado de falsedades'. Lo acusa de cometer 'el delito de pornografía infantil' al revelar el caso de una menor. Lo demandó por daño moral.",
     "264, 272-273, 282"),

    ("Paulo Díez Gargari", "Crítica", -3,
     "Lo califica de 'aberrante' y 'mercenario'. Lo acusa de difamar sin pruebas, de fraguar un plan con sobornos para revivir una concesión ferroviaria. Lo demandó por daño moral. Le dijo 'que se fuera a la chingada'. [Personaje marginal: aparece solo en las últimas páginas del libro.]",
     "277-282"),

    ("Manuel Bartlett", "Hostilidad", -4,
     "Dice que 'perjudicó muchísimo al presidente', que 'amenazó de muerte a mi papá'. Su gestión en CFE costó 300 millones de dólares en un arbitraje. 'No había buenos abogados' en su jurídico. La reforma eléctrica fue declarada inconstitucional.",
     "123-125, 212, 233-238"),

    ("Rutilio Escandón", "Hostilidad", -4,
     "Lo califica como 'el peor candidato que nos tocó: fue pavoroso'. 'No quería salir de su casa', tenía un discurso 'de región cuatro', 'se movía como un soldado de madera'. 'Un fiasco' total.",
     "94-96"),

    ("Hugo López-Gatell", "Hostilidad", -4,
     "Lo describe como 'personaje que hizo mucho daño', 'comunicador errático', 'funesto'. Lo acusa de impedir pruebas masivas, contradecirse sobre el tapabocas, adueñarse de la Cofepris. Dice que protegerlo fue una 'decisión pésima'. Subestimó la pandemia diciendo que 'no era una enfermedad grave', vendió mascarillas a China antes de necesitarlas, predijo máximo 60,000 muertos cuando fueron más de 600,000. Scherer le dedica un capítulo entero ('La pandemia') como ejemplo de la fórmula '90% lealtad, 10% capacidad'.",
     "158, 166-173, 197, 206, 271"),

    # --- NEGATIVOS (-3 a -2) ---
    ("Adán Augusto López Hernández", "Crítica", -3,
     "Dice que 'hubiera preferido mil veces trabajar con Octavio que con Adán'. Destruyó manglar sin permisos, impulsó eliminar licitaciones, 'gobernó a su antojo'. Como secretario de Gobernación, la relación con la Corte se fracturó bajo su gestión.",
     "87, 94-96, 226, 231-233, 250, 270, 276-277"),

    ("Elena Álvarez-Buylla", "Crítica", -3,
     "La califica como 'un desastre desde el principio', 'confrontada con la comunidad científica'. 'Engañó al presidente y nos metió en graves problemas'. 'Quiso meter a la cárcel a los investigadores'. Junto con Jesús Ramírez, convenció al presidente de prohibir el glifosato sin sustituto viable; el Conacyt bajo su dirección 'no fue capaz de producir el producto que sustituiría al glifosato. Ni siquiera se acercó.' También impulsó la demanda contra el maíz transgénico que se perdió ante el panel del T-MEC.",
     "132-133, 197-199"),

    ("Rosario Robles", "Crítica", -3,
     "Dice que 'fue una mujer apegada al dinero', que 'se descompuso' en el poder. 'Se involucró con personas que no le fueron totalmente leales'. También fue operadora electoral del PRI en el Estado de México.",
     "7, 21, 23, 36, 76"),

    ("Olga Sánchez Cordero", "Desconfianza", -2,
     "'Siempre había aspereza y fricción' entre ambos. La acusa de entregar una lista en su contra al presidente, de ser intermediaria en el caso Collado, y de alimentar embates mediáticos contra él. Apoyaba candidatos rivales de Zaldívar.",
     "108, 122, 137, 147, 153-154, 223-224, 237-238, 255-256, 284"),

    ("Miguel Ángel Yunes Linares", "Desconfianza", -2,
     "Lo presenta como adversario peligroso que mandó a Eva Cadena a calumniar a AMLO, que bloqueó camiones de campaña. Tenía una 'carpeta azul' con sus delitos pero 'era como de piedra'. Asociado con peligro físico.",
     "88, 98-99, 101-103"),

    ("Roberto Madrazo", "Crítica", -3,
     "Le 'ganó la elección' a AMLO en Tabasco 1994 mediante 'una elección de Estado', con recursos públicos, padrón de funcionarios y operadores locales.",
     "15"),

    ("Enrique Peña Nieto", "Crítica", -3,
     "'Echó a perder todo a partir de la Casa Blanca'. Lo increpó públicamente en un restaurante. Scherer lo confrontó de vuelta. 'Todo mundo empezó a desconfiar de él'. Asociado con operaciones electorales en Edomex.",
     "41, 54-57, 59-60, 64, 68, 76, 78-79, 103, 108, 126, 129, 140, 162-163"),

    ("Javier Quijano", "Desconfianza", -2,
     "Relación rota. Inicialmente amigo, pero lo acusó ante el presidente de intervenir en un asunto bancario. 'La relación está rota. Absolutamente.' Lo describe como 'abogado de derecha' que usaba cercanías como 'trampolín'.",
     "26, 37, 225, 269-271"),

    ("Jesús Ortega Martínez / 'los Chuchos'", "Crítica", -3,
     "'Andrés Manuel quería el poder; los Chuchos, el dinero.' Dice que 'decir que apoyaban a Andrés Manuel es una falacia'. Querían jubilarlo.",
     "36-37, 52, 64"),

    ("Luis Echeverría", "Crítica", -3,
     "Lo señala como responsable de la expulsión de su padre de Excélsior, junto con los mejores periodistas de su generación. Su comunicación se debía a su 'incapacidad para estarse tranquilo'.",
     "6, 159"),

    ("Felipe Calderón", "Crítica", -3,
     "Beneficiario de la campaña de 'peligro para México' financiada por empresarios. 'Sí lastimó' a AMLO. Su gobierno incorporó burocracia al Servicio Civil como obstáculo. Fue responsable de la desaparición de Luz y Fuerza.",
     "36, 40-41, 43-44, 56, 115, 132, 144, 154, 156, 159, 161-162, 173, 210-211"),

    ("Salomón Jara", "Crítica", -3,
     "'Un personaje diametralmente distinto a Susana' Harp. En las giras 'le mostraban cartulinas ofensivas con ratas horribles'. Tenía 'un gran pleito con la Sección 22'.",
     "86, 92, 99"),

    ("David Colmenares", "Desconfianza", -2,
     "Dice que 'no ha sido todo lo correcto que debió ser'. Lo acusa de firmar un convenio sospechoso con Nahle sobre auditorías de la refinería con observaciones incongruentes.",
     "194"),

    ("Ana Gabriela Guevara", "Desconfianza", -2,
     "'No le importaba el béisbol', era un obstáculo burocrático. Cuando 'decía que no, se frenaba la obra'. Había que invocar la autoridad del presidente para sortearla.",
     "129, 196-197"),

    ("Irma Eréndira Sandoval", "Desconfianza", -2,
     "Produjo un proyecto de Ley de Austeridad 'muy controvertido' con disposiciones 'casi violatorias y atentatorias de la ley'. Propuestas inviables que generarían corrupción.",
     "171, 191-192"),

    ("Claudio X. González (padre)", "Crítica", -3,
     "Lo confrontó directamente en el CCE: 'Usted sí pagó la campaña en 2006 contra López Obrador. Usted sí difamó e infamó a su familia, a su esposa, a sus hijos.'",
     "159-160"),

    ("Rocío Nahle", "Desconfianza", -2,
     "Candidata fuerte y sólida, pero destruyó manglar sin permisos junto con Adán Augusto. Bartlett la manipulaba. 'Todo el mérito, o el demérito, se lo llevó Rocío completito.'",
     "98-99, 102, 194, 232, 237-238"),

    ("Rosario Piedra Ibarra", "Crítica", -3,
     "La describe como 'inexperta en la materia' de derechos humanos. Bajo su presidencia la CNDH dio una respuesta absurda: consultó al acusado para determinar si había violación.",
     "258"),

    ("Enrique Molina", "Desconfianza", -2,
     "'Un empresario de un trato espantoso', 'un hombre de horrible trato'. Sin embargo, se negó a traicionarlo por principios éticos.",
     "27-28, 30"),

    ("Zedryk Raziel", "Crítica", -3,
     "Lo demandó por un 'texto difamatorio'.",
     "282"),

    ("Georgina Zerga", "Crítica", -3,
     "Demandada junto con Zedryk Raziel por un 'texto difamatorio'.",
     "282"),

    ("Marisela Morales", "Desconfianza", -2,
     "Operadora del desafuero, luego procuradora con Calderón, cónsule en Milán, después impulsada por Morena a la Corte. 'Contradicciones de la vida.'",
     "36"),

    ("Jesús Murillo Karam", "Desconfianza", -2,
     "'Fue duro conmigo' al saludarlo. Lo denunció junto con Lerdo de Tejada por una llamada grabada.",
     "56-57, 60"),

    ("Josefina Vázquez Mota", "Crítica", -3,
     "Dice que fue una 'candidatura débil'. El PAN con ella 'fue prácticamente acrítico con el PRI'. 'Se había desgastado mucho en la lucha interna'.",
     "54-55, 76-77, 281"),

    ("Erick Cisneros ('el Bola 8')", "Crítica", -3,
     "'Hizo todos los desmanes posibles' como secretario de Gobierno de Veracruz. Terminó peleado con Nahle y fuera del gabinete.",
     "99"),

    ("Laura Beristáin", "Crítica", -3,
     "'Lo más atrabancado del mundo'. Arrebató el micrófono para resaltar 'su propia figura. Lo demás no le importaba.'",
     "100"),

    ("Pedro Topete Vargas", "Crítica", -3,
     "Socio de Paulo Díez en el esquema de la concesión ferroviaria, en contubernio con funcionarios para beneficiarse 'a costa del erario'.",
     "278-279"),

    ("María Estela Ríos", "Desconfianza", -2,
     "Elegirla como consejera jurídica 'no fue una buena elección'. Permitió que Jesús Ramírez la asesorara en la defensa ante el T-MEC con resultados desastrosos.",
     "200, 233"),

    ("Rosalinda López", "Desconfianza", -2,
     "'Tenía un poder real sobre Rutilio. Lo supeditaba, lo manipulaba.' Junto con Adán Augusto, gobernaban Chiapas y Tabasco.",
     "96"),

    ("Vicente Fox", "Desconfianza", -2,
     "'Había perdido el respeto de la gente', con una 'pavorosa encuestitis'. Dio margen a Gil Díaz para su terrorismo fiscal. Su administración era 'a contentillo', sin forma ni consistencia.",
     "22, 24, 27-30, 36, 38, 40, 45, 112, 115, 159, 203, 263"),

    ("Alonso Ancira", "Desconfianza", -2,
     "Se negó a entregar Altos Hornos pese a estar paralizada. Vendió Agronitrogenados a Pemex 'a un precio elevadísimo' en lo que se sospechaba era un fraude.",
     "264-265"),

    ("Raquel Buenrostro", "Desconfianza", -2,
     "Se opuso al proyecto que Scherer respaldaba, generó desavenencias. La función de ratificar nombramientos generó 'errores y abusos'. Conflictos por compras consolidadas.",
     "117, 205-206, 239-242"),

    ("Clara Brugada", "Indiferencia", -1,
     "No la ataca directamente pero la presenta como beneficiaria de la operación clientelar de Jesús Ramírez, candidata impuesta en lugar del ganador de la encuesta, Omar García Harfuch.",
     "210-211, 219"),

    ("Norma Piña", "Indiferencia", -1,
     "La asocia con la fractura de la relación entre la Corte y el Ejecutivo en la segunda mitad del sexenio. Mención relativamente neutral.",
     "151, 226, 229"),

    ("Alejandro Moreno ('Alito')", "Desconfianza", -2,
     "'El más priista de los gobernadores.' 'Lo tenía todo controlado' en Campeche. 'Todos hablaron mal de Alito; había mucho temor.'",
     "87, 97, 101"),

    # --- NEUTRALES (0) ---
    ("Ernesto Zedillo", "Neutral", 0,
     "Mencionado contextualmente como presidente que no quería a Monreal como candidato, referencia temporal para reformas de seguridad.",
     "20, 27, 109, 140, 159, 179, 192"),

    ("Carlos Salinas de Gortari", "Neutral", 0,
     "Reconoce que 'a través de Solidaridad tuvo gran interlocución con la población', pero mencionado como parte del 'complot' según AMLO. Neutral.",
     "27, 29, 35, 159, 255"),

    ("Ricardo Anaya", "Neutral", 0,
     "Mencionado como candidato del PAN que 'enfocó su campaña en atacarse con Meade'. Sin opinión personal directa.",
     "76, 105"),

    ("José Antonio Meade", "Neutral", 0,
     "Lo cita respetuosamente reconociendo que si hubiera ganado 'el país nos habría reventado'. Neutral.",
     "105, 116, 164"),

    ("Alfonso Romo", "Neutral", 0,
     "Presentado como coordinador del equipo, figura pragmática del círculo interno. Sin juicio positivo o negativo fuerte.",
     "85, 105, 107-109, 112, 121, 123, 125, 224, 234"),

    ("Tatiana Clouthier", "Neutral", 0,
     "'Hizo un buen papel' en la Cámara, pero en un episodio lo marginó. Evaluación mixta que se cancela.",
     "105, 108, 122, 147, 197"),

    ("Luis María Aguilar", "Indiferencia", -1,
     "Su trato 'era bastante duro'. No le puso lugar en el desayuno de la Corte. La relación con AMLO 'terminó muy mal'.",
     "153, 222"),

    ("Cuauhtémoc Cárdenas", "Cordialidad", 1,
     "'Gran luchador social', más institucional que AMLO. La campaña de 1988 fue 'verdadera hazaña' y llenó el Zócalo. Se le puede considerar 'héroe'.",
     "7, 14, 19-20, 23, 39, 47"),

    ("Alfredo del Mazo", "Indiferencia", -1,
     "Candidato del PRI que ganó gracias a operativo gubernamental, no por mérito propio.",
     "76-79"),

    ("Miguel Ángel Mancera", "Desconfianza", -2,
     "'Cometió el error de apoyar a los candidatos del PRD contra Morena'. AMLO le dijo: 'De cuándo acá el perro manda al amo.' La seguridad se deterioró durante su gobierno.",
     "26, 37, 47, 68, 73-75"),

    ("Mario Delgado", "Cordialidad", 1,
     "'Tenía una muy buena relación con Mario Delgado'. Había crecido políticamente, quería ser jefe de Gobierno.",
     "73-74, 122, 147, 220, 235"),

    ("Santiago Nieto", "Neutral", 0,
     "Mencionado como director de la UIF que asistía al gabinete de seguridad. Sin juicio.",
     "136"),

    ("Bernardo Bátiz", "Indiferencia", -1,
     "AMLO 'siempre se quejó de que, por formalismos de Bátiz, Ahumada pudo salir adelante'. No logró judicializarlo.",
     "23, 35"),

    ("Esteban Moctezuma", "Cordialidad", 1,
     "'Muy buen tipo, muy razonable, trabajaba mucho en los diálogos.' Pero no tenía ánimo para empujar proyectos contra la burocracia. Junto con Scherer, plantearon al presidente conservar el instituto de evaluación educativa (INEE) con sus expertos en libros de texto y material didáctico, pero el presidente 'dijo que no era importante' y lo sustituyó por otro con gente de 'educación alternativa' que 'difícilmente eran expertos en educación'.",
     "107, 129-130, 196"),

    # --- POSITIVOS (1 a 2) ---
    ("Alfonso Durazo", "Aprecio", 2,
     "'Un tipo con muy buen discurso y gran capacidad organizativa'. 'Creo que sí fue buen secretario de Seguridad'. 'Hacía un gran papel' en el gabinete. Buena relación personal.",
     "9, 21-22, 29-30, 108, 114-116, 128, 131, 138, 143, 145, 176-177"),

    ("Dante Delgado", "Cordialidad", 1,
     "'Se había portado extraordinariamente bien con Andrés', 'nunca se negó a negociar, siempre ayudó'. Aunque reconoce su 'interés personal importante'.",
     "52-53, 147"),

    ("Alejandro Encinas", "Cordialidad", 1,
     "'Hizo un muy buen papel como suplente. Supo suavizar muchos temas.' Mantuvo distancia con Fox sin pelearse.",
     "38, 62"),

    ("Mauricio Kuri", "Aprecio", 2,
     "'Una pieza fundamental para que esto saliera adelante' en la negociación de la Guardia Nacional desde el PAN.",
     "146-147"),

    ("Claudia Ruiz Massieu", "Aprecio", 2,
     "'Muy buena negociadora: abogada asertiva, intuitiva y eficaz.' Reconoce su visión amplia del panorama.",
     "146-147"),

    ("Pedro Aspe", "Aprecio", 2,
     "'Siempre fue muy generoso conmigo. Me apapachaba mucho.' Lo presenta como hombre lúcido. 'A quien yo quiero mucho.'",
     "22, 27-29, 115, 263-264"),

    ("Delfina Gómez", "Cordialidad", 1,
     "'Muy sencilla, muy modesta', 'dúctil al cambio, participativa'. La llevaron de 14 a 34 puntos en encuestas.",
     "76-80"),

    ("Leonel Cota", "Aprecio", 2,
     "'Leal a sus principios', 'honesto', 'un hombre extraordinario'. 'Luchó hasta el último día por Andrés Manuel.'",
     "52"),

    ("Salvador Cienfuegos", "Aprecio", 2,
     "Lo defiende apasionadamente como inocente. 'No había una sola prueba contra Cienfuegos.' El expediente de la DEA era 'un invento intransitable'.",
     "113, 140, 181-189"),

    ("General Luis Cresencio Sandoval", "Aprecio", 2,
     "'Un hombre realmente profesional' con 'sensibilidad en derechos humanos'. Llevó 'información contundente' al presidente. Hizo 'un gran trabajo en la pandemia'.",
     "9, 113-114, 132, 136, 170, 177, 186, 206, 218, 239-240, 245"),

    ("Víctor Villalobos", "Aprecio", 2,
     "'Tenía muy claro lo que debía hacer', 'hombre muy eficaz'. Fue el primero en advertir que Álvarez-Buylla sería un desastre. Su consuegro. Defendió la inversión en sanidad animal argumentando que 'sin sanidad nos bloquearían la frontera con EE.UU.', pero el presidente 'decía que eso no se necesitaba, que solo era un gastadero de dinero'. 'Y eso fue, finalmente, lo que sucedió': se materializó el riesgo.",
     "121-122, 133, 197"),

    ("Santiago Levy", "Aprecio", 2,
     "'Un perfil social muy bueno', 'economista espléndido', creador de Progresa. 'Puede ayudarte muchísimo.' Lamenta que no aceptara ser secretario de Hacienda.",
     "109, 130"),

    ("Graciela Márquez", "Cordialidad", 1,
     "'Muy buena economista; además, es simpática y agradable. Hizo una buena labor.'",
     "130"),

    ("Jorge Alcocer", "Aprecio", 2,
     "'Un gran médico, una persona experta, un investigador.' Fue médico de cabecera de Rocío.",
     "130"),

    ("Jenaro Villamil", "Cordialidad", 1,
     "Su parte del libro 'fue muy buena'. 'Escribía sobre lo que ocurría en Televisa y otros medios.'",
     "49-50"),

    ("Juan Luis González Alcántara", "Aprecio", 2,
     "'Un abogado de primer orden', 'con orgullo, fue mi maestro'. Tuvo la valentía de advertirle directamente al presidente que votaría en contra.",
     "108, 148-149, 229"),

    ("Margarita Ríos Farjat", "Aprecio", 2,
     "'Persona muy sensata, muy buena abogada, muy competente.' 'Dura de carácter, fuerte.' Defendió su independencia.",
     "224, 229-230"),

    ("Yasmín Esquivel", "Cordialidad", 1,
     "'Recorrió todos los puestos' hasta la Corte. 'Hizo un trabajo espléndido.' Más receptiva a los pronunciamientos del Ejecutivo.",
     "229-230"),

    ("Loretta Ortiz", "Cordialidad", 1,
     "'Muchas veces escuchó bien las peticiones del Gobierno.' Actitud 'mucho menos beligerante que la de otros ministros.'",
     "230"),

    ("Verónica de Gives", "Aprecio", 2,
     "'Una magnífica abogada', 'una penalista excelente'. Lamenta que no pudiera asumir la Consejería Jurídica.",
     "232-233"),

    ("Carlos Urzúa", "Aprecio", 2,
     "Lo presenta con empatía como técnico que 'se quejaba amargamente' de las decisiones. 'Temblaba de coraje y de nervios.' Su renuncia fue consecuencia del conflicto.",
     "121, 130, 233-236"),

    ("Layda Sansores", "Cordialidad", 1,
     "'Muy querida y bien aceptada', con 'larga historia en Campeche'. 'Dicharachera, platicadora, con muy buen ambiente.'",
     "87, 97"),

    ("Manuel Velasco", "Aprecio", 2,
     "'Siempre generoso con Andrés en todos los sentidos. Incapaz de cometer una trapacería.' 'Estuvo a la altura' con los maestros. No intervino en la elección.",
     "86-87, 95-96, 103"),

    ("Eduardo Ramírez ('Lalo')", "Aprecio", 2,
     "'Candidato a senador buenísimo', 'construyó parte de la estructura', 'ayudó mucho', 'fue muy leal'. Hubiera sido mejor gobernador que Rutilio.",
     "95-96"),

    ("Rafael Marín Mollinedo", "Aprecio", 2,
     "'Un sujeto de primera, proveniente de la iniciativa privada.' 'Nos ayudó mucho a convencer a Mara Lezama.'",
     "88, 98"),

    ("Alfredo Higuera", "Aprecio", 2,
     "'De manera muy responsable' revisó el expediente Cienfuegos. 'Hizo un papel extraordinario en este caso.'",
     "187-188"),

    ("Alejandro Soberón", "Aprecio", 2,
     "'Hizo un trabajo maravilloso en la CDMX para que no faltaran camas' durante la pandemia. 'Empresario muy bueno.'",
     "170"),

    ("Héctor Grisi", "Cordialidad", 1,
     "Director global de Banco Santander. Le pidió consejo sobre la señora Botín. 'Quedó agradecido' y 'me lo correspondió.'",
     "269-270"),

    ("Julio Villarreal", "Aprecio", 2,
     "'Generoso con sus afectos, muy amable, siempre cariñoso conmigo.' Respondió desde Croacia a las 5 am para ayudar con financiamiento de campaña.",
     "59-60, 265"),

    ("Jorge Carrasco Araizaga", "Aprecio", 2,
     "Director de Proceso que confirmó ante Gertz que Scherer no intervenía en la revista. Defendió la independencia editorial.",
     "254"),

    ("Juan Antonio Pérez Simón", "Cordialidad", 1,
     "'Cariñoso como siempre', con gran colección de impresionistas. Destacó su generosidad al ofrecer prestar obra.",
     "132"),

    ("Humberto Artigas", "Aprecio", 2,
     "'Un hombre al que yo quiero muchísimo.' Arquitecto que aceptó ir a una reconciliación solo porque Scherer se lo pidió.",
     "271"),

    # --- MUY POSITIVOS (3 a 5) ---
    ("Ricardo Monreal", "Aprecio", 2,
     "'De los más experimentados en política', 'doctor en Derecho, aliado total'. Sin embargo, lo presionaba con 'recaditos para joderme' y lo acusó de ser parte de 'la tríada'. Relación ambivalente.",
     "9, 20, 73-74, 82, 84, 90-92, 99, 147"),

    ("Claudia Sheinbaum", "Admiración", 3,
     "'Líder que venía desde abajo', 'académica destacada', 'mujer sensible con los mejores sentimientos'. 'Aprende rápido, domina muchos temas.' 'Hizo un trabajo muy eficiente' en la pandemia. Le escribió una nota de confianza.",
     "23, 25, 46-47, 61, 73, 90-92, 121, 127, 137, 145, 151, 170-174, 177, 180, 207, 211, 217, 220, 237, 264, 282"),

    ("Marcelo Ebrard", "Admiración", 3,
     "'Un muy buen gobernador', administración 'ordenada, muy moderna'. 'No solo un político muy hábil, sino intelectualmente muy preparado.' 'Lo quiero y respeto mucho.' Gran mérito al declinar. Manejó bien el caso Cienfuegos.",
     "24, 34-35, 38, 40, 47-50, 55-56, 59, 62, 73-74, 82, 84, 116-117, 130, 170, 186, 188-190, 197, 220"),

    ("Rosa Icela Rodríguez", "Admiración", 3,
     "'Persona excepcional con tremendo feeling político.' 'La quiero de toda la vida. Nos decíamos hermanitos.' AMLO dice que era 'la persona que mejor conocía la ciudad'.",
     "24-25, 127"),

    ("Susana Harp", "Admiración", 3,
     "'Generosísima, buenísima, muy inteligente y muy culta; conocedora del ámbito indígena.' En las giras 'le gritaban cosas buenas, le aplaudían'.",
     "86, 92, 99"),

    ("César Yáñez", "Admiración", 3,
     "'Contribuía prodigiosamente', 'buen comunicador', 'persona muy generosa'. La decisión de apartarlo fue 'excesiva y lastimosa, porque si alguien era cercano a AMLO, ese era César'.",
     "39, 61, 67, 85, 104, 106, 116, 126-127"),

    ("Alberto Rojas ('Rojitas')", "Admiración", 3,
     "'Indispensable, imprescindible, y con nosotros siempre fue muy cariñoso, muy gentil, muy generoso.'",
     "104, 106, 116"),

    ("Laura Nieto ('Laurita')", "Admiración", 3,
     "'Una señora maravillosa que se consagró en cuerpo y alma a su causa.' 'Era una parte esencial.'",
     "39, 47-48, 106, 108"),

    ("Arturo Zaldívar", "Admiración", 3,
     "'Un gran abogado, el mejor amparista del país', 'hizo un trabajo espléndido'. 'El único ministro que se la jugó con López Obrador.' Le lastima lo que dicen de él. 'Trato magnífico.'",
     "4, 9, 151-155, 222-226, 229"),

    ("Carlos Montemayor", "Aprecio", 2,
     "'Un tipazo', 'ese magnífico escritor', 'mi amigo'. Tuvieron 'una relación maravillosa.'",
     "50, 57, 180"),

    ("Gonzalo López Beltrán", "Aprecio", 2,
     "'Lo quiero muchísimo', 'de una calidez, sencillez, humildad conmovedora'. 'Capacitador eficientísimo.' Aunque tenían tensiones por la visión de medios.",
     "89-90, 215"),

    ("Gabino Cué", "Cordialidad", 1,
     "AMLO lo describe como 'un tipo espléndido'. Scherer trabajó con él en Oaxaca. 'Platicamos largamente' y 'me fue muy bien'.",
     "62-64, 85"),

    ("Octavio Romero Oropeza", "Aprecio", 2,
     "'Soy amigo de Octavio.' 'Hubiera preferido mil veces trabajar con Octavio que con Adán.' Hay que exculparlo del tema de Dos Bocas.",
     "24, 87, 121, 176, 237-238"),

    ("Omar García Harfuch", "Admiración", 4,
     "'El personaje más valeroso que yo conozco.' 'De clase media, bien educado, de buenas formas.' 'Generosidad en el diálogo espléndida.' Ganó la encuesta por 20 puntos y 'debió haber sido jefe de Gobierno'. Lo conoce desde niño.",
     "8, 177-180, 210, 219-220"),

    ("Cuauhtémoc Blanco", "Indiferencia", -1,
     "'Pocas veces he visto a un hombre más popular.' Pero después lo describe con 'todos los inconvenientes que hoy le vemos'.",
     "110-111"),

    ("Beatriz Gutiérrez Müller", "Admiración", 3,
     "'Lindísima con mi papá', 'generosa, cariñosa, muy atenta', con 'agudeza intelectual'. 'Personaje fundamental en la vida de Andrés.' Contrataron juntos el seguro médico sin consentimiento de AMLO.",
     "12, 37, 39, 50-51, 62-63, 66-72, 102, 107, 131, 136, 154, 248-249, 270"),

    ("Heberto Castillo", "Admiración", 3,
     "'Una de las figuras fundacionales de la izquierda', 'hombre excepcional', 'hermano mayor' de su padre. 'Muy bueno, muy generoso.' Su renuncia en 1988 fue 'un acierto brutal'.",
     "5, 11, 14, 17-19, 69"),

    ("Javier García Paniagua", "Admiración", 4,
     "'Hombre extraordinario, simpático, divertido, generoso', con 'gran talento político, de una fuerza abrumadora'. 'Me llamaba hijo.' Fue su jefe desde 1978.",
     "24, 177-180"),

    ("Marcelino García Barragán", "Admiración", 3,
     "'De una modestia abrumadora.' Le encargó a don Julio cuidar a su hijo. Pidió solo 'un uniforme y un ataúd de madera' para su funeral.",
     "177-179"),

    ("María Sorté", "Aprecio", 2,
     "'Mujer extraordinaria, gran actriz, cariñosa, simpática y sensible.' Madre de Omar García Harfuch.",
     "179"),

    ("Mónica Pretelini", "Aprecio", 2,
     "'Mujer inteligente, culta, muy platicadora y desinhibida en lo político.' Su muerte 'me dolió mucho'.",
     "56"),

    ("Jorge Fernández Menéndez", "Admiración", 3,
     "'Amigo entrañable' que condujo las conversaciones para crear el libro. Comparten pasiones comunes.",
     "5-10"),

    ("Rocío Beltrán Medina", "Aprecio", 2,
     "'Mujer muy buena', 'la única persona que controlaba realmente a Andrés Manuel'. 'Alguien fundamental en su vida.'",
     "12, 17, 33"),

    ("Carmen Aristegui", "Cordialidad", 1,
     "'Hermosa relación' con su padre. Pero la critica por darle 'horas y horas' a Paulo Díez para difamarlo sin pruebas. Relación mixta con decepción.",
     "61, 216, 269, 281"),

    ("Alejandro Esquer", "Cordialidad", 1,
     "'Sonorense duro y rudo, pero buen hombre.' Encargado de toda la logística de AMLO.",
     "39, 67, 85, 106"),

    ("Julio Scherer García (padre)", "Devoción", 5,
     "Profunda admiración filial. 'El periodista más importante de México', 'hombre buenísimo'. 'Mi nombre es lo más valioso que poseo porque es el nombre intachable de mi padre.' Su legado es el eje moral de todo el libro.",
     "6, 11, 16-19, 30, 32-33, 49-51, 57-58, 62-63, 66-72, 89, 178, 215-216, 237, 251, 269, 281-283"),

    ("Andrés Manuel López Obrador", "Lealtad", 4,
     "'Compañero de travesía', 'mi hermano', 'cariño entrañable'. Lo admira como 'genio de la política' y líder transformador. Pero lo critica por dejarse manipular, no ser buen administrador, y la fórmula 90/10. 24 años de cercanísima amistad. 'La amistad verdadera sobrevive al odio y a los infundios.'",
     "Presente en prácticamente todas las páginas del libro"),

    ("Federico Arreola", "Cordialidad", 1,
     "Reconoce que hizo 'un trabajo muy importante con Andrés Manuel' y que 'se negó siempre' a aceptar dinero ilícito.",
     "161-162"),

    ("Miguel Ángel Granados Chapa", "Aprecio", 2,
     "'Gran periodista', parte del grupo de los mejores de su generación.",
     "6, 16, 269"),

    ("Rosario Ibarra de Piedra", "Aprecio", 2,
     "'Referente ético y moral' de AMLO, 'luchadora social desde los años setenta'. Su madre la llamaba 'mujer ejemplar'.",
     "258"),

    ("Guillermo Ortiz", "Cordialidad", 1,
     "Le dijo algo 'muy cariñoso' sobre cómo trataba a Gil Díaz, 'que es lo que se merece'.",
     "31"),

    ("Carlos Imaz", "Cordialidad", 1,
     "'Hombre muy valeroso.' Lo defendió legalmente. Critica que AMLO no hizo nada por él: 'una de las grandes injusticias'.",
     "25-26"),

    ("Armando Garza", "Cordialidad", 1,
     "'Un gran empresario'. Aprendió de él la lección de que el presidente tiene derecho a gobernar como decida.",
     "134"),

    ("Rafael Ramírez Heredia", "Cordialidad", 1,
     "Autor de 'El Rayo Macoy', 'muy divertido'. 'Le tenía mucho afecto a mi papá, y fue muy cariñoso conmigo.'",
     "269"),

    ("Manlio Fabio Beltrones", "Neutral", 0,
     "'Siempre muy amable y atento.' Lo alcanzó después del incidente con Peña para decirle que 'no tenía sentido'.",
     "56-57"),

    ("Emilio Gamboa Patrón", "Neutral", 0,
     "'Siempre muy amable y atento.'",
     "56-57"),

    ("Lázaro Cárdenas Batel", "Cordialidad", 1,
     "Acompañó gestiones por orden presidencial. Confía en que confirmará sus dichos.",
     "265"),

    ("Roberto Alcántara", "Cordialidad", 1,
     "Dueño de Viva Aerobús, con propuesta 'bastante razonable' para ayudar a los trabajadores de Interjet.",
     "266"),

    ("Ricardo Salinas Pliego", "Neutral", 0,
     "Presente en el festejo íntimo de la noche de la elección. Parte del consejo empresarial. Sin opinión directa.",
     "108, 163, 168"),

    ("Daniel Chávez", "Neutral", 0,
     "Presente en el festejo íntimo. Sin opinión directa.",
     "108"),

    ("Olegario Vázquez Aldir", "Neutral", 0,
     "Presente en el festejo íntimo. Parte del consejo empresarial. Sin opinión directa.",
     "107, 163, 168"),

    ("Bernardo Gómez", "Neutral", 0,
     "Presente en el festejo íntimo. Parte del consejo empresarial. Sin opinión directa.",
     "108, 163, 168"),

    ("Fidel Castro", "Neutral", 0,
     "Mención referencial: AMLO es 'admirador de Fidel Castro hasta el tuétano'.",
     "12"),

    ("Enrique Krauze", "Neutral", 0,
     "Mención puntual: calificó a AMLO como 'El mesías tropical'.",
     "12"),

    ("Octavio Paz", "Aprecio", 2,
     "'El gran Octavio Paz.' Citado en el epígrafe. Parte de los intelectuales expulsados de Excélsior.",
     "5-6, 202"),

    ("Carmen Lira", "Cordialidad", 1,
     "Directora de La Jornada. La reconciliación con su padre fue exitosa. 'Estoy seguro de que hoy lo recuerda con mucho cariño.'",
     "67, 70"),

    ("José María Pérez Gay", "Aprecio", 2,
     "'Hombre con experiencia diplomática, reconocido en medios intelectuales.' Lamenta su fallecimiento prematuro: 'Habría formado parte de su equipo más cercano.'",
     "35, 121"),

    ("Mauricio Vila", "Cordialidad", 1,
     "'Un candidatazo y luego un gran gobernador de Yucatán. Fue muy solidario con Andrés Manuel.'",
     "98, 101"),

    ("León Krauze", "Neutral", 0,
     "Sheinbaum reconoce que 'está haciendo un buen trabajo en un tema de migrantes'. Sin opinión de Scherer.",
     "217"),

    ("Luis Mandoki", "Aprecio", 2,
     "Colaborador cercano y amigo. Fue a dar un abrazo a AMLO la mañana de la derrota. Le dio espacio a su testimonio sobre el Charolazo.",
     "44, 48-49, 60-61, 129"),

    ("Elena Achar", "Desconfianza", -2,
     "'Aprovechó mi distracción para grabar secretamente un audio' que detonó el escándalo del Charolazo.",
     "61"),

    ("Enrique Álvarez del Castillo", "Cordialidad", 1,
     "'Asiduo a la pachanga' (lo dice 'con cariño'). Su frase sobre Proceso: 'No hay que salir ni para bien ni para mal.'",
     "58-59"),

    ("Graco Ramírez", "Neutral", 0,
     "Acompañante y amigo con quien compartía comida. Lo animó a reconciliarse con Peña.",
     "56-57"),

    ("Mara Lezama", "Cordialidad", 1,
     "Conductora de radio exitosa, persuadida por Rafael Marín. 'Ya estaba decidido que Mara sería la candidata y ganaría.'",
     "98, 100"),
]

# =============================================================================
# Calculate narrative weight from page references
# =============================================================================
import re

def count_pages(page_str):
    """Count individual pages from a page reference string like '27-32, 203, 263-264'."""
    if page_str.startswith("Presente"):
        return 300  # AMLO: present throughout
    total = 0
    for part in page_str.split(","):
        part = part.strip()
        match = re.match(r"(\d+)-(\d+)", part)
        if match:
            total += int(match.group(2)) - int(match.group(1)) + 1
        elif re.match(r"\d+", part):
            total += 1
    return total

def get_weight(page_str):
    """Classify narrative weight: Alta (15+), Media (5-14), Baja (1-4)."""
    n = count_pages(page_str)
    if n >= 15:
        return "Alta"
    elif n >= 5:
        return "Media"
    else:
        return "Baja"

# =============================================================================
# Sort by score (ascending: most hated first)
# =============================================================================
personas.sort(key=lambda x: (x[2], x[0]))

# =============================================================================
# Create Excel workbook
# =============================================================================
wb = Workbook()

# --- Sheet 1: Ranking ---
ws = wb.active
ws.title = "Ranking"

# Styles
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

red_dark = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
red_med = PatternFill(start_color="FF7777", end_color="FF7777", fill_type="solid")
red_light = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
orange_light = PatternFill(start_color="FFD699", end_color="FFD699", fill_type="solid")
yellow_light = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
neutral_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
green_light = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
green_med = PatternFill(start_color="85E085", end_color="85E085", fill_type="solid")
green_dark = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
green_very_dark = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

body_font = Font(name="Calibri", size=11)
body_font_white = Font(name="Calibri", size=11, color="FFFFFF")
wrap_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top")

# Headers
headers = ["Ranking", "Nombre", "Opinión", "Puntuación", "Peso narrativo", "Descripción", "Páginas"]
col_widths = [9, 30, 14, 13, 16, 75, 22]

col_letters = ["A", "B", "C", "D", "E", "F", "G"]
for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[col_letters[col_idx - 1]].width = width

# Data rows
for row_idx, (name, opinion, score, desc, pages) in enumerate(personas, 2):
    ranking = row_idx - 1
    weight = get_weight(pages)

    ws.cell(row=row_idx, column=1, value=ranking).alignment = center_align
    ws.cell(row=row_idx, column=2, value=name)
    ws.cell(row=row_idx, column=3, value=opinion).alignment = center_align
    ws.cell(row=row_idx, column=4, value=score).alignment = center_align
    ws.cell(row=row_idx, column=5, value=weight).alignment = center_align
    ws.cell(row=row_idx, column=6, value=desc).alignment = wrap_align
    ws.cell(row=row_idx, column=7, value=pages).alignment = wrap_align

    # Color coding
    if score <= -4:
        fill = red_dark
        font = body_font_white
    elif score == -3:
        fill = red_med
        font = body_font
    elif score == -2:
        fill = red_light
        font = body_font
    elif score == -1:
        fill = orange_light
        font = body_font
    elif score == 0:
        fill = neutral_fill
        font = body_font
    elif score == 1:
        fill = green_light
        font = body_font
    elif score == 2:
        fill = green_med
        font = body_font
    elif score == 3:
        fill = green_dark
        font = body_font
    elif score >= 4:
        fill = green_very_dark
        font = body_font_white
    else:
        fill = neutral_fill
        font = body_font

    for col_idx in range(1, 8):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = thin_border

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:G{len(personas) + 1}"

# --- Sheet 2: Metodología ---
ws2 = wb.create_sheet("Metodología")
ws2.column_dimensions["A"].width = 100

methodology = [
    "METODOLOGÍA DE ANÁLISIS",
    "",
    "Fuente: 'Ni venganza ni perdón' de Julio Scherer Ibarra (319 páginas)",
    "Análisis realizado por Claude Code (Opus 4.6) — Febrero 2026",
    "",
    "PROCESO:",
    "1. Se extrajo el texto completo del PDF página por página usando pdfplumber.",
    "2. Se dividió en 15 bloques de ~20 páginas cada uno.",
    "3. Se analizaron los 15 bloques en paralelo, identificando TODAS las personas reales mencionadas.",
    "4. Se consolidaron variantes del mismo nombre (ej: 'AMLO', 'Andrés Manuel', 'López Obrador' = misma persona).",
    "5. Se clasificó la opinión de Scherer basándose ESTRICTAMENTE en el texto del libro.",
    "6. Se asignó puntuación y se generó este archivo.",
    "",
    "ESCALA DE OPINIÓN:",
    "-5 = Desprecio (lo detesta profundamente, persecución activa, lenguaje visceral)",
    "-4 = Hostilidad (antagonismo fuerte, lo considera muy dañino)",
    "-3 = Crítica (juicio negativo claro, desaprobación explícita)",
    "-2 = Desconfianza (recelo, malestar, fricciones)",
    "-1 = Indiferencia (mención con tono ligeramente desfavorable)",
    " 0 = Neutral (mención factual sin juicio de valor)",
    " 1 = Cordialidad (trato amable, reconocimiento moderado)",
    " 2 = Aprecio (valoración positiva clara, elogios específicos)",
    " 3 = Admiración (elogios enfáticos, expresiones de respeto profundo)",
    " 4 = Lealtad (vínculo personal fuerte, defensa apasionada)",
    " 5 = Devoción (amor filial, eje moral del libro)",
    "",
    "CRITERIOS DE CLASIFICACIÓN:",
    "• Solo se clasifican opiniones que Scherer expresa directamente en el texto.",
    "• Si menciona a alguien sin emitir juicio, se clasifica como Neutral (0).",
    "• Se priorizó la opinión PREDOMINANTE cuando hay ambivalencia.",
    "• Las citas textuales se parafrasean respetando el sentido original.",
    "• Las páginas refieren al PDF original.",
    "",
    "CÓDIGO DE COLORES:",
    "• Rojo oscuro: Puntuación -5 a -4 (Desprecio / Hostilidad)",
    "• Rojo medio: Puntuación -3 (Crítica)",
    "• Rojo claro: Puntuación -2 (Desconfianza)",
    "• Naranja: Puntuación -1 (Indiferencia negativa)",
    "• Gris: Puntuación 0 (Neutral)",
    "• Verde claro: Puntuación 1 (Cordialidad)",
    "• Verde medio: Puntuación 2 (Aprecio)",
    "• Verde oscuro: Puntuación 3 (Admiración)",
    "• Verde muy oscuro: Puntuación 4-5 (Lealtad / Devoción)",
    "",
    "PESO NARRATIVO (columna nueva en v2):",
    "• Alta: 15 o más páginas dedicadas al personaje (protagonistas del relato)",
    "• Media: 5-14 páginas (personajes recurrentes)",
    "• Baja: 1-4 páginas (menciones puntuales o episódicas)",
    "• Este indicador complementa la puntuación de opinión y permite distinguir",
    "  entre antagonistas centrales (ej. Ramírez Cuevas, Gertz) y personajes",
    "  marginales (ej. Paulo Díez, que solo aparece en las últimas páginas).",
    "",
    "TEMAS DE POLÍTICA PÚBLICA MENCIONADOS EN EL LIBRO:",
    "• Sanidad animal: Scherer relata que Villalobos (Agricultura) defendió la inversión",
    "  en sanidad animal, pero AMLO 'decía que eso no se necesitaba, que solo era",
    "  un gastadero de dinero'. Villalobos advertía que sin sanidad bloquearían la",
    "  frontera con EE.UU., 'y eso fue, finalmente, lo que sucedió' (pp. 121-122).",
    "• Instituto de evaluación educativa (INEE): Scherer y Moctezuma plantearon",
    "  conservar el instituto con expertos que 'analizaba los libros de texto, el",
    "  material didáctico'. El presidente 'dijo que no era importante' y lo sustituyó",
    "  por otro con gente que 'difícilmente eran expertos en educación' (pp. 129-130).",
    "• Pandemia COVID-19: Capítulo extenso (pp. 169-176). Scherer critica duramente",
    "  a López-Gatell por subestimar la enfermedad, impedir pruebas masivas, y",
    "  contradecirse sobre el tapabocas. Reconoce que los militares y Sheinbaum",
    "  hicieron buen trabajo frente al 'desastre' federal. Califica la decisión de",
    "  proteger a López-Gatell como 'pésima' (fórmula 90% lealtad / 10% capacidad).",
    "• Glifosato y maíz transgénico: Álvarez-Buylla (Conacyt), Jesús Ramírez y",
    "  López-Gatell convencieron al presidente de prohibir el glifosato sin sustituto.",
    "  El Conacyt 'no fue capaz de producir el producto sustituto. Ni siquiera se",
    "  acercó.' Ramírez asesoró la defensa del T-MEC sin formación legal, provocando",
    "  'un revés histórico para la soberanía alimentaria' (pp. 197-200).",
    "",
    "NOTAS:",
    f"• Total de personas identificadas: {len(personas)}",
    "• Personas con opinión negativa (< 0): " + str(sum(1 for p in personas if p[2] < 0)),
    "• Personas neutrales (= 0): " + str(sum(1 for p in personas if p[2] == 0)),
    "• Personas con opinión positiva (> 0): " + str(sum(1 for p in personas if p[2] > 0)),
    "• El personaje más extensamente tratado es Andrés Manuel López Obrador,",
    "  presente en prácticamente todas las páginas del libro.",
    "• Julio Scherer García (padre del autor) es el referente moral del libro.",
    "",
    "CAMBIOS EN v2 (atendiendo observaciones):",
    "• Se agregó columna 'Peso narrativo' (Alta/Media/Baja) para distinguir",
    "  protagonistas de personajes episódicos.",
    "• Paulo Díez Gargari recalibrado de -5 a -3: personaje marginal (solo pp. 277-282).",
    "• Se enriquecieron descripciones de antagonistas centrales (Ramírez Cuevas,",
    "  Gertz Manero) para reflejar su presencia a lo largo de todo el libro.",
    "• Se vincularon temas de política pública (sanidad animal, INEE, pandemia,",
    "  glifosato) a los personajes correspondientes en sus descripciones.",
]

title_font = Font(name="Calibri", bold=True, size=14)
subtitle_font = Font(name="Calibri", bold=True, size=11)
normal_font = Font(name="Calibri", size=11)

for row_idx, line in enumerate(methodology, 1):
    cell = ws2.cell(row=row_idx, column=1, value=line)
    if row_idx == 1:
        cell.font = title_font
    elif line.endswith(":") and line == line.upper():
        cell.font = subtitle_font
    elif line.startswith(("-5", "-4", "-3", "-2", "-1", " 0", " 1", " 2", " 3", " 4", " 5")):
        cell.font = normal_font
    else:
        cell.font = normal_font
    cell.alignment = Alignment(wrap_text=True)

# Save
output_path = os.path.join(OUTPUT_DIR, "amigos_o_rivales_JSI.xlsx")
wb.save(output_path)
print(f"Excel saved to: {output_path}")
print(f"Total personas: {len(personas)}")
print(f"Negativos: {sum(1 for p in personas if p[2] < 0)}")
print(f"Neutrales: {sum(1 for p in personas if p[2] == 0)}")
print(f"Positivos: {sum(1 for p in personas if p[2] > 0)}")
